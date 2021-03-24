import openpyxl


def getRowCount(filePath, sheetName):
    workbook = openpyxl.load_workbook(filePath)

    sheet = workbook[sheetName]

    return sheet.max_row


def getColumnCount(filePath, sheetName):
    workbook = openpyxl.load_workbook(filePath)

    sheet = workbook[sheetName]

    return sheet.max_column


def readData(filePath, sheetName, rownum, columno):
    workbook = openpyxl.load_workbook(filePath)

    sheet = workbook[sheetName]

    return sheet.cell(row=rownum, column=columno).value


def writeData(filePath, sheetName, rownum, columnm, data):
    workbook = openpyxl.load_workbook(filePath)

    sheet = workbook[sheetName]

    sheet.cell(row=rownum, column=columnm, ).value = data

    workbook.save(filePath)
